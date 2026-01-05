import os
import sys
import signal
import logging
import subprocess
from threading import Lock, Thread
import time
import sqlite3
import json
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
from flask import Flask, render_template, send_file, request, Response, redirect, url_for, session, jsonify
from flask_socketio import SocketIO, emit
import pandas as pd
from collections import defaultdict, deque
import glob
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
import re
import tenacity
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from PIL import Image

# Setup logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("ping_debug.log"),
        logging.StreamHandler()
    ]
)

# Configuration
app = Flask(__name__, template_folder='templates')
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', 'secret!')
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading', ping_timeout=120, ping_interval=30)
results_lock = Lock()
results = []
running = True
PING_INTERVAL = 10
XLSX_FILE = "sm_ips.xlsx"
RETENTION_DAYS = 7
LATENCY_THRESHOLD = 800
DEGRADED_LOSS_THRESHOLD = 10
MAX_WORKERS = 50
BATCH_SIZE = 20
LAST_PRUNE = None
LAST_XLSX_LOAD = None
CACHED_DF = None
ALERT_LOG = deque(maxlen=1000)
ALERT_QUEUE = deque(maxlen=1000)
LOG_DIR = "logs"
CURRENT_LOG_FILE = None
LAST_LOG_DATE = None
LAST_LOG_HOUR = None
ALERT_UPDATE_INTERVAL = 3.0
ALERT_TIMESTAMP_COUNTER = 0
HISTORY_CACHE = {}
SUMMARY_CACHE = {}
status_cache = {}
previous_status_cache = {}
downtime_cache = {}
uptime_cache = {}
alert_cache = {}
alert_thread_running = True

# Ensure log directory exists
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)
    logging.info(f"Created log directory: {LOG_DIR}")

def format_duration(seconds):
    """Convert seconds to HH:MM:SS format."""
    if not isinstance(seconds, (int, float)) or seconds < 0:
        return "N/A"
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"

def get_log_filename():
    now = datetime.now()
    return os.path.join(LOG_DIR, f"alerts_{now.strftime('%Y-%m-%d_%H')}.jsonl")

def rotate_log_file():
    global CURRENT_LOG_FILE, LAST_LOG_DATE, LAST_LOG_HOUR
    now = datetime.now()
    if LAST_LOG_DATE != now.date() or LAST_LOG_HOUR != now.hour:
        CURRENT_LOG_FILE = get_log_filename()
        LAST_LOG_DATE = now.date()
        LAST_LOG_HOUR = now.hour
        logging.info(f"Rotated log file to {CURRENT_LOG_FILE}")
    elif os.path.exists(CURRENT_LOG_FILE) and os.path.getsize(CURRENT_LOG_FILE) > 10 * 1024 * 1024:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        os.rename(CURRENT_LOG_FILE, f"{CURRENT_LOG_FILE}.{timestamp}")
        CURRENT_LOG_FILE = get_log_filename()
        logging.info(f"Rotated log file due to size limit to {CURRENT_LOG_FILE}")

def append_to_log_file(alert):
    rotate_log_file()
    try:
        with open(CURRENT_LOG_FILE, 'a', encoding='utf-8') as f:
            json.dump(alert, f)
            f.write('\n')
    except Exception as e:
        logging.error(f"Failed to append to log file: {str(e)}")

@tenacity.retry(stop=tenacity.stop_after_attempt(3), wait=tenacity.wait_fixed(1))
def init_db():
    try:
        conn = sqlite3.connect('ping_history.db', check_same_thread=False)
        conn.execute('''CREATE TABLE IF NOT EXISTS history 
                        (timestamp TEXT, sm_ip TEXT, status TEXT, latency REAL)''')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_sm_ip ON history (sm_ip)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_timestamp ON history (timestamp)')
        conn.commit()
        conn.close()
        logging.info("Database initialized")
    except Exception as e:
        logging.error(f"Database init failed: {str(e)}")
        raise

init_db()

def prune_old_records():
    global LAST_PRUNE
    try:
        now = datetime.now()
        if LAST_PRUNE and now - LAST_PRUNE < timedelta(hours=1):
            return
        cutoff = (now - timedelta(days=RETENTION_DAYS)).strftime('%Y-%m-%d %H:%M:%S')
        with sqlite3.connect('ping_history.db') as conn:
            conn.execute("DELETE FROM history WHERE timestamp < ?", (cutoff,))
            conn.commit()
        LAST_PRUNE = now
        logging.info("Pruned old records")
    except Exception as e:
        logging.error(f"Prune error: {str(e)}")

def validate_ip(ip):
    ip_pattern = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
    return bool(ip and re.match(ip_pattern, ip))

def ping_ip(ip, timeout=0.2):
    try:
        if ip == "192.168.138.141" and os.getenv("TEST_MODE"):
            logging.debug(f"Simulating {ip}: Reachable")
            return "Reachable", 33.25

        cached = status_cache.get(ip)
        if cached and datetime.now() - cached["time"] < timedelta(seconds=PING_INTERVAL * 5):
            logging.debug(f"Using cached status for {ip}: {cached['status']}")
            return cached["status"], cached["latency"]
        
        cmd = ['ping', '-c', '2', '-W', str(timeout), ip]
        logging.debug(f"Pinging {ip}")
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=3)
        
        stdout = result.stdout.lower()
        if result.returncode != 0 or "100% packet loss" in stdout or "destination host unreachable" in stdout:
            logging.warning(f"{ip} Down")
            status_cache[ip] = {"status": "Down", "latency": None, "time": datetime.now()}
            return "Down", None
        
        loss_percent = 100
        latency = None
        for line in stdout.splitlines():
            if "packet loss" in line:
                loss_percent = int(line.split(",")[2].split("%")[0].strip())
            if "rtt min/avg/max/mdev" in line:
                latency = float(line.split('=')[1].split('/')[1])
        
        if loss_percent >= DEGRADED_LOSS_THRESHOLD or (latency and latency > LATENCY_THRESHOLD):
            status = "Degraded"
        else:
            status = "Reachable"
        
        status_cache[ip] = {"status": status, "latency": latency, "time": datetime.now()}
        logging.debug(f"{ip} {status}, latency: {latency or 'N/A'}")
        return status, latency
    except subprocess.TimeoutExpired:
        logging.error(f"Ping {ip} timed out")
        status_cache[ip] = {"status": "Down", "latency": None, "time": datetime.now()}
        return "Down", None
    except Exception as e:
        logging.error(f"Ping {ip} error: {str(e)}")
        status_cache[ip] = {"status": "Down", "latency": None, "time": datetime.now()}
        return "Down", None

def ping_all_ips(ips):
    start = time.time()
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        batches = [ips[i:i + BATCH_SIZE] for i in range(0, len(ips), BATCH_SIZE)]
        ping_results = []
        for batch in batches:
            ping_results.extend(executor.map(ping_ip, batch))
    duration = time.time() - start
    logging.info(f"Pinging {len(ips)} IPs took {duration:.2f}s")
    return ping_results

@tenacity.retry(stop=tenacity.stop_after_attempt(3), wait=tenacity.wait_fixed(1))
def log_to_db(entries):
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            conn.executemany("INSERT INTO history VALUES (?, ?, ?, ?)", entries)
            conn.commit()
        logging.debug(f"Logged {len(entries)} entries to DB")
    except Exception as e:
        logging.error(f"DB log error: {str(e)}")
        raise

def get_downtime_since(sm_ip, current_status):
    if current_status != "Down":
        return "N/A"
    cached = downtime_cache.get(sm_ip)
    if cached and datetime.now() - cached["time"] < timedelta(minutes=2):
        return cached["value"]
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cutoff = (datetime.now() - timedelta(hours=48)).strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute("""
                SELECT timestamp FROM history 
                WHERE sm_ip = ? AND status = 'Down' AND timestamp >= ?
                ORDER BY timestamp DESC LIMIT 1
            """, (sm_ip, cutoff))
            record = cursor.fetchone()
            if not record:
                logging.warning(f"No recent Down record for {sm_ip} in last 48 hours")
                return "Unknown"
            result = record[0]
            downtime_cache[sm_ip] = {"value": result, "time": datetime.now()}
            return result
    except Exception as e:
        logging.error(f"Downtime {sm_ip}: {str(e)}")
        return "Unknown"

def get_uptime_since(sm_ip, current_status):
    if current_status != "Reachable":
        return "N/A"
    cached = uptime_cache.get(sm_ip)
    if cached and datetime.now() - cached["time"] < timedelta(minutes=2):
        return cached["value"]
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT timestamp FROM history 
                WHERE sm_ip = ? AND status = 'Reachable' 
                ORDER BY timestamp DESC LIMIT 1
            """, (sm_ip,))
            record = cursor.fetchone()
            if not record:
                return "Unknown"
            result = record[0]
            uptime_cache[sm_ip] = {"value": result, "time": datetime.now()}
            return result
    except Exception as e:
        logging.error(f"Uptime {sm_ip}: {str(e)}")
        return "Unknown"

def get_previous_status_from_db(sm_ip):
    try:
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT status FROM history 
                WHERE sm_ip = ? 
                ORDER BY timestamp DESC LIMIT 1
            """, (sm_ip,))
            record = cursor.fetchone()
            return record[0] if record else None
    except Exception as e:
        logging.error(f"DB previous status {sm_ip}: {str(e)}")
        return None

def update_alert_log():
    global alert_thread_running
    while running and alert_thread_running:
        try:
            if ALERT_QUEUE:
                alert = ALERT_QUEUE.popleft()
                if alert['entry']['status'] not in ['Error', 'Timeout']:
                    ALERT_LOG.append(alert)
                    append_to_log_file(alert)
                    cutoff_time = datetime.now() - timedelta(hours=1)
                    recent_alerts = [
                        a for a in ALERT_LOG
                        if datetime.strptime(a['time'].split('.')[0], '%Y-%m-%d %H:%M:%S') >= cutoff_time
                        and a['entry']['status'] in ['Down', 'Degraded', 'Reachable']
                    ]
                    socketio.emit('alert_update', {'alerts': recent_alerts})
                    logging.debug(f"Emitted {len(recent_alerts)} alerts for {alert['time']} - {alert['entry']['status']} for {alert['entry']['ip']}")
            time.sleep(ALERT_UPDATE_INTERVAL)
        except Exception as e:
            logging.error(f"Alert log update error: {str(e)}")
            alert_thread_running = False
            time.sleep(ALERT_UPDATE_INTERVAL)

def monitor_alert_thread():
    global alert_thread_running
    while running:
        if not alert_thread_running:
            logging.warning("Alert thread failed, restarting...")
            alert_thread_running = True
            alert_thread = Thread(target=update_alert_log, name="AlertUpdateThread", daemon=True)
            alert_thread.start()
            logging.info("Alert thread restarted")
        time.sleep(10)

@tenacity.retry(stop=tenacity.stop_after_attempt(3), wait=tenacity.wait_fixed(1))
def update_ping_status():
    global results, running, CACHED_DF, LAST_XLSX_LOAD, ALERT_TIMESTAMP_COUNTER
    socketio.emit('update_status', {
        'results': [],
        'pop_summary': {},
        'alerts': [],
        'analysis': {}
    })
    while running:
        start_time = time.time()
        logging.info("Starting ping cycle")
        try:
            prune_old_records()
            
            now = datetime.now()
            if CACHED_DF is None or LAST_XLSX_LOAD is None or now - LAST_XLSX_LOAD > timedelta(minutes=5):
                try:
                    if not os.path.exists(XLSX_FILE):
                        logging.error(f"{XLSX_FILE} missing")
                        raise FileNotFoundError(f"{XLSX_FILE} not found")
                    CACHED_DF = pd.read_excel(XLSX_FILE, engine='openpyxl')
                    LAST_XLSX_LOAD = now
                    logging.debug(f"Loaded XLSX with {len(CACHED_DF)} rows")
                except Exception as e:
                    logging.error(f"XLSX load error: {str(e)}")
                    CACHED_DF = pd.DataFrame(columns=['AP Name', 'AP IP', 'CID', 'SM IP', 'Device Name', 'Location'])
                    socketio.emit('update_status', {
                        'results': [{"AP Name": "N/A", "AP IP": "N/A", "CID": "N/A", "SM IP": "N/A", "Device Name": f"XLSX load failed: {str(e)}", "Location": "N/A", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
                        'pop_summary': {},
                        'analysis': {}
                    })
                    time.sleep(5)
                    continue
            
            df = CACHED_DF
            if 'SM IP' not in df.columns or 'Location' not in df.columns or df.empty:
                error_msg = f"Invalid XLSX: Missing required columns or empty. Columns: {df.columns.tolist()}"
                logging.error(error_msg)
                socketio.emit('update_status', {
                    'results': [{"AP Name": "N/A", "AP IP": "N/A", "CID": "N/A", "SM IP": "N/A", "Device Name": error_msg, "Location": "N/A", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
                    'pop_summary': {},
                    'analysis': {}
                })
                time.sleep(5)
                continue
            
            # Debug: Log column names and sample data
            logging.info(f"Excel columns: {df.columns.tolist()}")
            if not df.empty:
                logging.info(f"Sample row: {df.iloc[0].to_dict()}")
            
            valid_locations = df['Location'].dropna().astype(str).str.strip()
            if valid_locations.empty:
                error_msg = "No valid Location values in XLSX"
                logging.error(error_msg)
                socketio.emit('update_status', {
                    'results': [{"AP Name": "N/A", "AP IP": "N/A", "CID": "N/A", "SM IP": "N/A", "Device Name": error_msg, "Location": "N/A", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
                    'pop_summary': {},
                    'analysis': {}
                })
                time.sleep(5)
                continue
            
            ip_info = {
                str(row['SM IP']): {
                    'org_name': str(row.get('Device Name', 'N/A')),
                    'location': str(row.get('Location', 'Unknown')),
                    'ap_name': str(row.get('AP Name', 'N/A')),
                    'ap_ip': str(row.get('AP IP', 'N/A')),
                    'cid': str(row.get('CID', 'N/A'))
                } for row in df.to_dict('records') if pd.notna(row.get('SM IP'))
            }
            
            # Debug: Log a few sample entries to see CID values
            sample_ips = list(ip_info.keys())[:3]
            for sample_ip in sample_ips:
                logging.info(f"Sample IP {sample_ip}: {ip_info[sample_ip]}")
            
            ips = list(ip_info.keys())
            logging.info(f"Pinging {len(ips)} IPs")

            ping_results = ping_all_ips(ips)
            ping_duration = time.time() - start_time
            ip_to_result = dict(zip(ips, ping_results))

            new_results = []
            pop_counts = defaultdict(lambda: {'Reachable': 0, 'Degraded': 0, 'Down': 0})
            db_entries = []
            reachable_counts = defaultdict(int)
            ip_stats = defaultdict(lambda: {
                'uptime': 0,
                'downtime': 0,
                'degraded_time': 0,
                'days': set(),
                'down_count': 0,
                'up_count': 0,
                'down_events': [],
                'up_events': [],
                'degraded_events': []
            })

            for sm_ip in ips:
                status, latency = ip_to_result.get(sm_ip, ("Unknown", None))
                if status in ['Error', 'Timeout']:
                    continue
                logging.debug(f"Ping result for {sm_ip}: {status}, latency: {latency or 'N/A'}")
                
                last_status = previous_status_cache.get(sm_ip, {}).get("status")
                if last_status is None:
                    last_status = get_previous_status_from_db(sm_ip)
                
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                db_entries.append((timestamp, sm_ip, status, latency))
                
                info = ip_info[sm_ip]
                downtime_since = get_downtime_since(sm_ip, status)
                uptime_since = get_uptime_since(sm_ip, status)
                is_long_term = False
                if downtime_since not in ["N/A", "Unknown"]:
                    try:
                        downtime_start = datetime.strptime(downtime_since, '%Y-%m-%d %H:%M:%S')
                        duration = (datetime.now() - downtime_start).total_seconds()
                        is_long_term = duration >= 24 * 3600
                        logging.debug(f"Long-term check for {sm_ip}: downtime_since={downtime_since}, is_long_term={is_long_term}")
                    except Exception as e:
                        logging.error(f"Error parsing downtime for {sm_ip}: {str(e)}")
                
                result = {
                    "AP Name": info['ap_name'],
                    "AP IP": info['ap_ip'],
                    "CID": info['cid'],
                    "SM IP": sm_ip,
                    "Device Name": info['org_name'],
                    "Location": info['location'],
                    "Status": status,
                    "Latency": f"{latency:.2f} ms" if latency is not None else "N/A",
                    "Downtime Since": downtime_since
                }
                new_results.append(result)
                pop_counts[info['location']][status] += 1
                
                ip_stats[sm_ip]['days'].add(datetime.now().strftime('%Y-%m-%d'))
                if status == 'Down':
                    ip_stats[sm_ip]['downtime'] += PING_INTERVAL
                    if last_status != 'Down':
                        ip_stats[sm_ip]['down_count'] += 1
                        ip_stats[sm_ip]['down_events'].append(timestamp)
                elif status == 'Reachable':
                    ip_stats[sm_ip]['uptime'] += PING_INTERVAL
                    if last_status != 'Reachable':
                        ip_stats[sm_ip]['up_count'] += 1
                        ip_stats[sm_ip]['up_events'].append(timestamp)
                elif status == 'Degraded':
                    ip_stats[sm_ip]['degraded_time'] += PING_INTERVAL
                    if last_status != 'Degraded':
                        ip_stats[sm_ip]['degraded_events'].append(timestamp)
                
                alert_entry = {
                    "ip": sm_ip,
                    "location": info.get('location', 'N/A'),
                    "org_name": info.get('org_name', 'N/A'),
                    "ap_name": info.get('ap_name', 'N/A'),
                    "ap_ip": info.get('ap_ip', 'N/A'),
                    "cid": info.get('cid', 'N/A'),
                    "status": status,
                    "downtime_since": downtime_since,
                    "uptime_since": uptime_since,
                    "latency": latency,
                    "long_term": is_long_term,
                    "high_latency": latency is not None and latency > LATENCY_THRESHOLD
                }
                should_generate = False
                if status in ["Down", "Degraded"] or is_long_term:
                    should_generate = True
                elif status == "Reachable" and last_status in ["Down", "Degraded"]:
                    should_generate = True
                    reachable_counts[sm_ip] += 1
                
                cached_alert = alert_cache.get(sm_ip, {})
                if should_generate and (
                    not cached_alert or
                    cached_alert.get("status") != status or
                    (is_long_term and not cached_alert.get("long_term"))
                ):
                    ALERT_TIMESTAMP_COUNTER += 1
                    timestamp = f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.{ALERT_TIMESTAMP_COUNTER:06d}"
                    alert = {
                        "time": timestamp,
                        "entry": alert_entry
                    }
                    ALERT_LOG.append(alert)
                    ALERT_QUEUE.append(alert)
                    append_to_log_file(alert)
                    alert_cache[sm_ip] = {"status": status, "long_term": is_long_term, "time": datetime.now()}
                    logging.debug(f"Generated alert: {status} for {sm_ip}, long_term: {is_long_term}, previous status: {last_status or 'None'}")
                    if is_long_term:
                        alert_cache[sm_ip] = {"status": status, "long_term": False, "time": datetime.now()}
                
                previous_status_cache[sm_ip] = {"status": status, "latency": latency, "time": datetime.now()}

            if db_entries:
                log_to_db(db_entries)

            new_results.sort(key=lambda x: (
                0 if x['Status'] == 'Down' else 1 if x['Status'] == 'Degraded' else 2))
            with results_lock:
                results = new_results
            
            ip_daily_stats = {}
            for sm_ip in ip_stats:
                days = len(ip_stats[sm_ip]['days']) or 1
                ip_daily_stats[sm_ip] = {
                    'uptime': format_duration(ip_stats[sm_ip]['uptime']),
                    'downtime': format_duration(ip_stats[sm_ip]['downtime']),
                    'degraded_time': format_duration(ip_stats[sm_ip]['degraded_time']),
                    'down_count': ip_stats[sm_ip]['down_count'],
                    'up_count': ip_stats[sm_ip]['up_count'],
                    'down_events': ip_stats[sm_ip]['down_events'],
                    'up_events': ip_stats[sm_ip]['up_events'],
                    'degraded_events': ip_stats[sm_ip]['degraded_events'],
                    'location': ip_info.get(sm_ip, {}).get('location', 'Unknown')
                }
            
            logging.info(f"Emitting update_status: {len(new_results)} results, {len(pop_counts)} locations in pop_summary")
            
            # Debug: Log first few results to see what's being sent
            if new_results:
                for i, result in enumerate(new_results[:3]):
                    logging.info(f"Result {i}: {result}")
            
            socketio.emit('update_status', {
                'results': new_results,
                'pop_summary': dict(pop_counts),
                'ip_stats': ip_daily_stats
            })
            logging.info(f"Cycle completed in {time.time() - start_time:.2f}s")
        except Exception as e:
            logging.exception(f"Cycle error: {str(e)}")
            socketio.emit('update_status', {
                'results': [{"AP Name": "N/A", "AP IP": "N/A", "CID": "N/A", "SM IP": "N/A", "Device Name": f"Cycle failed: {str(e)}", "Location": "N/A", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
                'pop_summary': {},
                'ip_stats': {}
            })
        finally:
            elapsed = time.time() - start_time
            sleep_time = max(0, PING_INTERVAL - elapsed)
            logging.debug(f"Sleeping {sleep_time:.2f}s")
            time.sleep(sleep_time)

def start_periodic_update():
    ping_thread = Thread(target=update_ping_status, name="PingUpdateThread", daemon=True)
    alert_thread = Thread(target=update_alert_log, name="AlertUpdateThread", daemon=True)
    monitor_thread = Thread(target=monitor_alert_thread, name="MonitorAlertThread", daemon=True)
    ping_thread.start()
    alert_thread.start()
    monitor_thread.start()
    logging.info("Update and monitor threads started")

def signal_handler(sig, frame):
    global running
    running = False
    logging.info("Shutting down")
    sys.exit(0)

signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

# SocketIO Events
@socketio.on('refresh_now')
def handle_refresh_now():
    logging.info("Received refresh_now request")
    try:
        with results_lock:
            socketio.emit('update_status', {
                'results': results,
                'pop_summary': SUMMARY_CACHE.get('pop_summary', {}),
                'ip_stats': SUMMARY_CACHE.get('ip_stats', {})
            })
    except Exception as e:
        logging.error(f"Refresh now error: {str(e)}")
        socketio.emit('update_status', {
            'results': [{"AP Name": "N/A", "AP IP": "N/A", "CID": "N/A", "SM IP": "N/A", "Device Name": f"Refresh failed: {str(e)}", "Location": "N/A", "Status": "Error", "Latency": "N/A", "Downtime Since": "N/A"}],
            'pop_summary': {},
            'ip_stats': {}
        })

@socketio.on('refresh_alerts')
def handle_refresh_alerts():
    logging.info("Received refresh_alerts request")
    try:
        cutoff_time = datetime.now() - timedelta(hours=1)
        recent_alerts = [
            a for a in ALERT_LOG
            if datetime.strptime(a['time'].split('.')[0], '%Y-%m-%d %H:%M:%S') >= cutoff_time
            and a['entry']['status'] in ['Down', 'Degraded', 'Reachable']
        ]
        socketio.emit('refresh_alerts', {'alerts': recent_alerts})
        logging.debug(f"Emitted {len(recent_alerts)} alerts for refresh")
    except Exception as e:
        logging.error(f"Refresh alerts error: {str(e)}")
        socketio.emit('refresh_alerts', {'error': str(e)})

# Admin Routes
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        admin_username = os.getenv('ADMIN_USERNAME', 'admin')
        admin_password = os.getenv('ADMIN_PASSWORD', 'admin123')
        if username == admin_username and password == admin_password:
            session['logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            return render_template('admin_login.html', error="Invalid credentials")
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('logged_in', None)
    return redirect(url_for('admin_login'))

@app.route('/admin')
def admin_dashboard():
    if not session.get('logged_in'):
        return redirect(url_for('admin_login'))
    try:
        df = pd.read_excel(XLSX_FILE, engine='openpyxl')
        locations = sorted(set(df['Location'].dropna().astype(str)))
        records = df.to_dict('records')
        return render_template('admin.html', records=records, locations=locations)
    except Exception as e:
        logging.error(f"Admin dashboard error: {str(e)}")
        return render_template('admin.html', records=[], locations=[], error=str(e))

@app.route('/admin/add_entry', methods=['POST'])
def admin_add_entry():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    global CACHED_DF, LAST_XLSX_LOAD
    try:
        data = request.json
        sm_ip = data.get('sm_ip')
        if not sm_ip or not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid or missing SM IP'}), 400
        ap_ip = data.get('ap_ip', 'N/A')
        if ap_ip != 'N/A' and not validate_ip(ap_ip):
            return jsonify({'error': 'Invalid AP IP'}), 400
        cid = data.get('cid', 'N/A')
        location = data.get('location')
        if not location:
            return jsonify({'error': 'Location is required'}), 400
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            if sm_ip in CACHED_DF['SM IP'].values:
                return jsonify({'error': 'SM IP already exists'}), 400
            new_entry = {
                'AP Name': data.get('ap_name', 'N/A'),
                'AP IP': ap_ip,
                'CID': cid,
                'SM IP': sm_ip,
                'Device Name': data.get('org_name', 'N/A'),
                'Location': location
            }
            CACHED_DF = pd.concat([CACHED_DF, pd.DataFrame([new_entry])], ignore_index=True)
            CACHED_DF.to_excel(XLSX_FILE, index=False, engine='openpyxl')
            LAST_XLSX_LOAD = None
            logging.info(f"Added new entry for SM IP {sm_ip}")
            return jsonify({'success': 'Entry added successfully'})
    except Exception as e:
        logging.error(f"Add entry error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/update_entry', methods=['POST'])
def admin_update_entry():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    global CACHED_DF, LAST_XLSX_LOAD
    try:
        data = request.json
        sm_ip = data.get('sm_ip')
        if not sm_ip or not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid or missing SM IP'}), 400
        ap_ip = data.get('ap_ip', 'N/A')
        if ap_ip != 'N/A' and not validate_ip(ap_ip):
            return jsonify({'error': 'Invalid AP IP'}), 400
        mac_pattern = r'^([0-9A-Fa-f]{2}:){5}[0-9A-Fa-f]{2}$'
        ap_mac = data.get('ap_mac', 'N/A')
        sm_mac = data.get('sm_mac', 'N/A')
        if ap_mac != 'N/A' and not re.match(mac_pattern, ap_mac):
            return jsonify({'error': 'Invalid AP MAC Address'}), 400
        if sm_mac != 'N/A' and not re.match(mac_pattern, sm_mac):
            return jsonify({'error': 'Invalid SM MAC Address'}), 400
        location = data.get('location')
        if not location:
            return jsonify({'error': 'Location is required'}), 400
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            if sm_ip not in CACHED_DF['SM IP'].values:
                return jsonify({'error': 'SM IP not found'}), 404
            update_data = {
                'AP Name': data.get('ap_name', CACHED_DF.loc[CACHED_DF['SM IP'] == sm_ip, 'AP Name'].iloc[0]),
                'AP IP': ap_ip,
                'AP MAC Address': ap_mac,
                'SM IP': sm_ip,
                'Device Name': data.get('org_name', CACHED_DF.loc[CACHED_DF['SM IP'] == sm_ip, 'Device Name'].iloc[0]),
                'SM MAC Address': sm_mac,
                'Location': location
            }
            CACHED_DF.loc[CACHED_DF['SM IP'] == sm_ip] = pd.DataFrame([update_data])
            CACHED_DF.to_excel(XLSX_FILE, index=False, engine='openpyxl')
            LAST_XLSX_LOAD = None
            logging.info(f"Updated entry for SM IP {sm_ip}")
            return jsonify({'success': 'Entry updated successfully'})
    except Exception as e:
        logging.error(f"Update entry error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/delete_entry', methods=['POST'])
def admin_delete_entry():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    global CACHED_DF, LAST_XLSX_LOAD
    try:
        data = request.json
        sm_ip = data.get('sm_ip')
        if not sm_ip or not validate_ip(sm_ip):
            return jsonify({'error': 'Invalid or missing SM IP'}), 400
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            if sm_ip not in CACHED_DF['SM IP'].values:
                return jsonify({'error': 'SM IP not found'}), 404
            CACHED_DF = CACHED_DF[CACHED_DF['SM IP'] != sm_ip]
            CACHED_DF.to_excel(XLSX_FILE, index=False, engine='openpyxl')
            LAST_XLSX_LOAD = None
            logging.info(f"Deleted entry for SM IP {sm_ip}")
            return jsonify({'success': 'Entry deleted successfully'})
    except Exception as e:
        logging.error(f"Delete entry error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/location_downtime')
def location_downtime():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        location_filter = request.args.get('location_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None or 'SM IP' not in CACHED_DF.columns or 'Location' not in CACHED_DF.columns:
                return jsonify({'error': 'Dataframe not initialized or missing required columns'}), 500
            ip_to_location = {str(row['SM IP']): str(row['Location']) for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
            all_locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        downtime_data = defaultdict(lambda: {
            'total_downtime': 0,
            'downtime_count': 0,
            'down_ips': set(),
            'last_downtime': None,
            'uptime': 'N/A'
        })
        ip_stats = defaultdict(lambda: {
            'uptime': 0,
            'downtime': 0,
            'degraded_time': 0,
            'days': set(),
            'down_count': 0,
            'up_count': 0,
            'down_events': [],
            'up_events': [],
            'degraded_events': []
        })
        for i, (sm_ip, status, timestamp) in enumerate(records):
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            ts = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            day = ts.strftime('%Y-%m-%d')
            next_status = None
            next_ts = None
            if i + 1 < len(records) and records[i + 1][0] == sm_ip:
                next_status = records[i + 1][1]
                next_ts = datetime.strptime(records[i + 1][2], '%Y-%m-%d %H:%M:%S')
            duration = PING_INTERVAL
            if next_status and next_ts:
                duration = (next_ts - ts).total_seconds()
            if status == 'Down':
                downtime_data[location]['total_downtime'] += duration
                downtime_data[location]['down_ips'].add(sm_ip)
                downtime_data[location]['downtime_count'] += 1
                if not downtime_data[location]['last_downtime'] or ts > datetime.strptime(downtime_data[location]['last_downtime'], '%Y-%m-%d %H:%M:%S'):
                    downtime_data[location]['last_downtime'] = timestamp
                ip_stats[sm_ip]['downtime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Down':
                    ip_stats[sm_ip]['down_count'] += 1
                    ip_stats[sm_ip]['down_events'].append(timestamp)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
            elif status == 'Reachable':
                ip_stats[sm_ip]['uptime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Reachable':
                    ip_stats[sm_ip]['up_count'] += 1
                    ip_stats[sm_ip]['up_events'].append(timestamp)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
            elif status == 'Degraded':
                ip_stats[sm_ip]['degraded_time'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Degraded':
                    ip_stats[sm_ip]['degraded_events'].append(timestamp)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
        for location in downtime_data:
            down_ips = downtime_data[location]['down_ips']
            if down_ips:
                cursor.execute("""
                    SELECT sm_ip, timestamp
                    FROM history
                    WHERE sm_ip IN ({}) AND status = 'Reachable' AND timestamp >= ? AND timestamp < ?
                    ORDER BY timestamp DESC
                """.format(','.join(['?'] * len(down_ips))), list(down_ips) + [start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')])
                uptime_records = cursor.fetchall()
                latest_uptime = {}
                for sm_ip, timestamp in uptime_records:
                    if sm_ip not in latest_uptime:
                        latest_uptime[sm_ip] = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                if latest_uptime:
                    latest = max(latest_uptime.values())
                    duration = (datetime.now() - latest).total_seconds()
                    downtime_data[location]['uptime'] = format_duration(duration)
        result = {
            'locations': [],
            'ip_stats': {},
            'chart_data': {
                'bar': {
                    'labels': [],
                    'data': []
                }
            }
        }
        locations_to_process = [location_filter] if location_filter else all_locations
        for location in locations_to_process:
            data = downtime_data.get(location, {
                'total_downtime': 0,
                'downtime_count': 0,
                'down_ips': set(),
                'last_downtime': None,
                'uptime': 'N/A'
            })
            avg_downtime = format_duration(data['total_downtime'] / data['downtime_count']) if data['downtime_count'] > 0 else "N/A"
            result['locations'].append({
                'location': location,
                'total_downtime': format_duration(data['total_downtime']),
                'avg_downtime': avg_downtime,
                'down_ip_count': len(data['down_ips']),
                'down_ip_link': f"/admin/location_ips?location={location}&start_date={start_date}&end_date={end_date}",
                'last_downtime': data['last_downtime'],
                'uptime': data['uptime']
            })
            if data['total_downtime'] > 0:
                result['chart_data']['bar']['labels'].append(location)
                result['chart_data']['bar']['data'].append(data['total_downtime'])
        sorted_ips = sorted(
            ip_stats.items(),
            key=lambda x: x[1]['downtime'],
            reverse=True
        )
        worst_ips = sorted_ips[:10]
        best_ips = sorted_ips[-10:][::-1] if len(sorted_ips) >= 10 else sorted_ips[::-1]
        for sm_ip, stats in worst_ips + best_ips:
            result['ip_stats'][sm_ip] = {
                'location': stats['location'],
                'uptime': format_duration(stats['uptime']),
                'downtime': format_duration(stats['downtime']),
                'degraded_time': format_duration(stats['degraded_time']),
                'down_count': stats['down_count'],
                'up_count': stats['up_count'],
                'down_events': stats['down_events'],
                'up_events': stats['up_events'],
                'degraded_events': stats['degraded_events']
            }
        logging.info(f"Retrieved downtime data for {len(result['locations'])} locations")
        return jsonify(result)
    except Exception as e:
        logging.error(f"Location downtime error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/export_location_downtime')
def export_location_downtime():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        location_filter = request.args.get('location_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None or 'SM IP' not in CACHED_DF.columns or 'Location' not in CACHED_DF.columns:
                return jsonify({'error': 'Dataframe not initialized or missing required columns'}), 500
            ip_to_location = {str(row['SM IP']): str(row['Location']) for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
            all_locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp, latency
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        downtime_data = defaultdict(lambda: {
            'total_downtime': 0,
            'downtime_count': 0,
            'down_ips': set(),
            'last_downtime': None,
            'uptime': 'N/A'
        })
        ip_stats = defaultdict(lambda: {
            'uptime': 0,
            'downtime': 0,
            'degraded_time': 0,
            'days': set(),
            'down_count': 0,
            'up_count': 0,
            'down_events': [],
            'up_events': [],
            'degraded_events': [],
            'down_durations': [],
            'up_durations': [],
            'degraded_durations': []
        })
        for i, (sm_ip, status, timestamp, latency) in enumerate(records):
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            ts = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            day = ts.strftime('%Y-%m-%d')
            next_status = None
            next_ts = None
            if i + 1 < len(records) and records[i + 1][0] == sm_ip:
                next_status = records[i + 1][1]
                next_ts = datetime.strptime(records[i + 1][2], '%Y-%m-%d %H:%M:%S')
            duration = PING_INTERVAL
            if next_status and next_ts:
                duration = (next_ts - ts).total_seconds()
            if status == 'Down':
                downtime_data[location]['total_downtime'] += duration
                downtime_data[location]['down_ips'].add(sm_ip)
                downtime_data[location]['downtime_count'] += 1
                if not downtime_data[location]['last_downtime'] or ts > datetime.strptime(downtime_data[location]['last_downtime'], '%Y-%m-%d %H:%M:%S'):
                    downtime_data[location]['last_downtime'] = timestamp
                ip_stats[sm_ip]['downtime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Down':
                    ip_stats[sm_ip]['down_count'] += 1
                    ip_stats[sm_ip]['down_events'].append(timestamp)
                ip_stats[sm_ip]['down_durations'].append(duration)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
            elif status == 'Reachable':
                ip_stats[sm_ip]['uptime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Reachable':
                    ip_stats[sm_ip]['up_count'] += 1
                    ip_stats[sm_ip]['up_events'].append(timestamp)
                ip_stats[sm_ip]['up_durations'].append(duration)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
            elif status == 'Degraded':
                ip_stats[sm_ip]['degraded_time'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Degraded':
                    ip_stats[sm_ip]['degraded_events'].append(timestamp)
                ip_stats[sm_ip]['degraded_durations'].append(duration)
                ip_stats[sm_ip]['days'].add(day)
                ip_stats[sm_ip]['location'] = location
        for location in downtime_data:
            down_ips = downtime_data[location]['down_ips']
            if down_ips:
                cursor.execute("""
                    SELECT sm_ip, timestamp
                    FROM history
                    WHERE sm_ip IN ({}) AND status = 'Reachable' AND timestamp >= ? AND timestamp < ?
                    ORDER BY timestamp DESC
                """.format(','.join(['?'] * len(down_ips))), list(down_ips) + [start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')])
                uptime_records = cursor.fetchall()
                latest_uptime = {}
                for sm_ip, timestamp in uptime_records:
                    if sm_ip not in latest_uptime:
                        latest_uptime[sm_ip] = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
                if latest_uptime:
                    latest = max(latest_uptime.values())
                    duration = (datetime.now() - latest).total_seconds()
                    downtime_data[location]['uptime'] = format_duration(duration)
        wb = Workbook()
        ws = wb.active
        ws.title = "Location Downtime Report"
        headers = ['Location', 'Total Downtime (HH:MM:SS)', 'Avg Downtime (HH:MM:SS)', 'Down IP Count', 'Last Downtime', 'Uptime Since Last Down (HH:MM:SS)']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        locations_to_process = [location_filter] if location_filter else all_locations
        for location in locations_to_process:
            data = downtime_data.get(location, {
                'total_downtime': 0,
                'downtime_count': 0,
                'down_ips': set(),
                'last_downtime': None,
                'uptime': 'N/A'
            })
            avg_downtime = format_duration(data['total_downtime'] / data['downtime_count']) if data['downtime_count'] > 0 else "N/A"
            ws.append([
                location,
                format_duration(data['total_downtime']),
                avg_downtime,
                len(data['down_ips']),
                data['last_downtime'] or 'N/A',
                data['uptime']
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        ws = wb.create_sheet("IP Stats")
        headers = [
            'SM IP', 'Location', 'Times Down', 'Times Up', 'Avg Downtime (HH:MM:SS)', 
            'Avg Uptime (HH:MM:SS)', 'Total Downtime (HH:MM:SS)', 'Total Uptime (HH:MM:SS)', 
            'Total Degraded Time (HH:MM:SS)', 'Down Events', 'Up Events', 'Degraded Events'
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sorted_ips = sorted(
            ip_stats.items(),
            key=lambda x: x[1]['downtime'],
            reverse=True
        )
        for sm_ip, stats in sorted_ips:
            avg_downtime = format_duration(sum(stats['down_durations']) / len(stats['down_durations'])) if stats['down_durations'] else "N/A"
            avg_uptime = format_duration(sum(stats['up_durations']) / len(stats['up_durations'])) if stats['up_durations'] else "N/A"
            ws.append([
                sm_ip,
                stats['location'],
                stats['down_count'],
                stats['up_count'],
                avg_downtime,
                avg_uptime,
                format_duration(stats['downtime']),
                format_duration(stats['uptime']),
                format_duration(stats['degraded_time']),
                '; '.join(stats['down_events']),
                '; '.join(stats['up_events']),
                '; '.join(stats['degraded_events'])
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        ws = wb.create_sheet("Detailed Logs")
        headers = ['Timestamp', 'SM IP', 'Status', 'Latency (ms)', 'Location']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for sm_ip, status, timestamp, latency in records:
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            ws.append([
                timestamp,
                sm_ip,
                status,
                f"{latency:.2f}" if latency is not None else 'N/A',
                location
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logging.info(f"Exported location downtime report for {start_date} to {end_date}")
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"location_downtime_{start_date}_to_{end_date}.xlsx"
        )
    except Exception as e:
        logging.error(f"Export location downtime error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/export_log')
def export_log():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        cutoff_time = datetime.now() - timedelta(hours=1)
        recent_alerts = [
            a for a in ALERT_LOG
            if datetime.strptime(a['time'].split('.')[0], '%Y-%m-%d %H:%M:%S') >= cutoff_time
            and a['entry']['status'] in ['Down', 'Degraded', 'Reachable']
        ]
        wb = Workbook()
        ws = wb.active
        ws.title = "Alert Log"
        headers = ['Timestamp', 'SM IP', 'Status', 'Location', 'Device Name', 'AP Name', 'AP IP', 'AP MAC', 'SM MAC', 'Latency', 'Downtime Since', 'Uptime Since', 'Long Term']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for alert in recent_alerts:
            entry = alert['entry']
            ws.append([
                alert['time'],
                entry['ip'],
                entry['status'],
                entry['location'],
                entry['org_name'],
                entry['ap_name'],
                entry['ap_ip'],
                entry['ap_mac'],
                entry['sm_mac'],
                f"{entry['latency']:.2f}" if entry['latency'] is not None else 'N/A',
                entry['downtime_since'],
                entry['uptime_since'],
                'Yes' if entry['long_term'] else 'No'
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logging.info("Exported alert log")
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"alert_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except Exception as e:
        logging.error(f"Export log error: {str(e)}")
        return jsonify({'error': str(e)}), 500
















@app.route('/export_ip_history')
def export_ip_history():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        ip_input = request.args.get('ip', '')
        date = request.args.get('date')
        hour = request.args.get('hour')
        format_type = request.args.get('format', 'xlsx')
        sm_ips = [ip.strip() for ip in ip_input.split(',') if validate_ip(ip.strip())]
        if not sm_ips:
            return jsonify({'error': 'Invalid or missing SM IP(s)'}), 400
        query_params = []
        query_conditions = ["sm_ip IN ({})".format(','.join(['?'] * len(sm_ips)))]
        query_params.extend(sm_ips)
        if date:
            try:
                date_obj = datetime.strptime(date, '%Y-%m-%d')
                start_time = date_obj.strftime('%Y-%m-%d 00:00:00')
                end_time = (date_obj + timedelta(days=1)).strftime('%Y-%m-%d 00:00:00')
                query_conditions.append("timestamp >= ? AND timestamp < ?")
                query_params.extend([start_time, end_time])
            except ValueError:
                return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        if hour:
            try:
                hour_int = int(hour)
                if not (0 <= hour_int <= 23):
                    return jsonify({'error': 'Invalid hour. Must be 00-23'}), 400
                if date:
                    start_time = datetime.strptime(f"{date} {hour}:00:00", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                    end_time = datetime.strptime(f"{date} {hour}:59:59", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                    query_conditions.pop()
                    query_conditions.append("timestamp >= ? AND timestamp < ?")
                    query_params[-2:] = [start_time, end_time]
                else:
                    return jsonify({'error': 'Date must be specified if hour is provided'}), 400
            except ValueError:
                return jsonify({'error': 'Invalid hour format'}), 400
        query = f"""
            SELECT timestamp, sm_ip, status, latency
            FROM history
            WHERE {' AND '.join(query_conditions)}
            ORDER BY timestamp
        """
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute(query, query_params)
            records = cursor.fetchall()
        if not records:
            return jsonify({'error': f'No history found for IP(s) {ip_input}'}), 404
        if format_type.lower() != 'xlsx':
            return jsonify({'error': 'Unsupported format. Use xlsx'}), 400
        ip_stats = defaultdict(lambda: {
            'uptime': 0,
            'downtime': 0,
            'degraded_time': 0,
            'down_count': 0,
            'up_count': 0,
            'down_events': [],
            'up_events': [],
            'degraded_events': [],
            'down_durations': [],
            'up_durations': [],
            'degraded_durations': []
        })
        for i, (timestamp, sm_ip, status, latency) in enumerate(records):
            ts = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            next_status = None
            next_ts = None
            if i + 1 < len(records) and records[i + 1][1] == sm_ip:
                next_status = records[i + 1][2]
                next_ts = datetime.strptime(records[i + 1][0], '%Y-%m-%d %H:%M:%S')
            duration = PING_INTERVAL
            if next_status and next_ts:
                duration = (next_ts - ts).total_seconds()
            if status == 'Down':
                ip_stats[sm_ip]['downtime'] += duration
                if i == 0 or records[i-1][1] != sm_ip or records[i-1][2] != 'Down':
                    ip_stats[sm_ip]['down_count'] += 1
                    ip_stats[sm_ip]['down_events'].append(timestamp)
                ip_stats[sm_ip]['down_durations'].append(duration)
            elif status == 'Reachable':
                ip_stats[sm_ip]['uptime'] += duration
                if i == 0 or records[i-1][1] != sm_ip or records[i-1][2] != 'Reachable':
                    ip_stats[sm_ip]['up_count'] += 1
                    ip_stats[sm_ip]['up_events'].append(timestamp)
                ip_stats[sm_ip]['up_durations'].append(duration)
            elif status == 'Degraded':
                ip_stats[sm_ip]['degraded_time'] += duration
                if i == 0 or records[i-1][1] != sm_ip or records[i-1][2] != 'Degraded':
                    ip_stats[sm_ip]['degraded_events'].append(timestamp)
                ip_stats[sm_ip]['degraded_durations'].append(duration)
        wb = Workbook()
        ws = wb.active
        ws.title = "IP History"
        headers = [
            'Timestamp', 'SM IP', 'Status', 'Latency (ms)', 'Downtime Since', 
            'Uptime Since', 'Times Down', 'Times Up', 'Avg Downtime (HH:MM:SS)', 
            'Avg Uptime (HH:MM:SS)', 'Total Downtime (HH:MM:SS)', 'Total Uptime (HH:MM:SS)', 
            'Total Degraded Time (HH:MM:SS)', 'Down Events', 'Up Events', 'Degraded Events'
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ip_to_location = {str(row['SM IP']): str(row['Location']) for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
        for sm_ip in sm_ips:
            stats = ip_stats[sm_ip]
            avg_downtime = format_duration(sum(stats['down_durations']) / len(stats['down_durations'])) if stats['down_durations'] else "N/A"
            avg_uptime = format_duration(sum(stats['up_durations']) / len(stats['up_durations'])) if stats['up_durations'] else "N/A"
            for timestamp, _, status, latency in [(r[0], r[1], r[2], r[3]) for r in records if r[1] == sm_ip]:
                ws.append([
                    timestamp,
                    sm_ip,
                    status,
                    f"{latency:.2f}" if latency is not None else 'N/A',
                    get_downtime_since(sm_ip, status),
                    get_uptime_since(sm_ip, status),
                    stats['down_count'],
                    stats['up_count'],
                    avg_downtime,
                    avg_uptime,
                    format_duration(stats['downtime']),
                    format_duration(stats['uptime']),
                    format_duration(stats['degraded_time']),
                    '; '.join(stats['down_events']),
                    '; '.join(stats['up_events']),
                    '; '.join(stats['degraded_events'])
                ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logging.info(f"Exported IP history for {ip_input}")
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"ip_history_{'_'.join(sm_ips)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
    except Exception as e:
        logging.error(f"Export IP history error: {str(e)}")
        return jsonify({'error': str(e)}), 500





@app.route('/get_logs')
def get_logs():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        ip_input = request.args.get('ip', '')
        date = request.args.get('date')
        hour = request.args.get('hour')
        sm_ips = [ip.strip() for ip in ip_input.split(',') if validate_ip(ip.strip())]
        if not sm_ips:
            return jsonify({'error': 'Invalid or missing SM IP(s)'}), 400
        cache_key = f"{ip_input}_{date}_{hour}"
        if cache_key in HISTORY_CACHE:
            cached = HISTORY_CACHE[cache_key]
            if datetime.now() - cached['time'] < timedelta(minutes=5):
                logging.debug(f"Returning cached logs for {cache_key}")
                return jsonify(cached['data'])
        query_params = []
        query_conditions = ["sm_ip IN ({})".format(','.join(['?'] * len(sm_ips)))]
        query_params.extend(sm_ips)
        if date:
            try:
                date_obj = datetime.strptime(date, '%Y-%m-%d')
                start_time = date_obj.strftime('%Y-%m-%d 00:00:00')
                end_time = (date_obj + timedelta(days=1)).strftime('%Y-%m-%d 00:00:00')
                query_conditions.append("timestamp >= ? AND timestamp < ?")
                query_params.extend([start_time, end_time])
            except ValueError:
                return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        if hour:
            try:
                hour_int = int(hour)
                if not (0 <= hour_int <= 23):
                    return jsonify({'error': 'Invalid hour. Must be 00-23'}), 400
                if date:
                    start_time = datetime.strptime(f"{date} {hour}:00:00", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                    end_time = datetime.strptime(f"{date} {hour}:59:59", '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
                    query_conditions.pop()
                    query_conditions.append("timestamp >= ? AND timestamp < ?")
                    query_params[-2:] = [start_time, end_time]
                else:
                    return jsonify({'error': 'Date must be specified if hour is provided'}), 400
            except ValueError:
                return jsonify({'error': 'Invalid hour format'}), 400
        query = f"""
            SELECT timestamp, sm_ip, status, latency
            FROM history
            WHERE {' AND '.join(query_conditions)}
            ORDER BY timestamp
        """
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            cursor.execute(query, query_params)
            records = cursor.fetchall()
        logs = []
        ip_info = {}
        with results_lock:
            if CACHED_DF is not None:
                ip_info = {
                    str(row['SM IP']): {
                        'org_name': str(row.get('Device Name', 'N/A')),
                        'location': str(row.get('Location', 'Unknown')),
                        'ap_name': str(row.get('AP Name', 'N/A')),
                        'ap_ip': str(row.get('AP IP', 'N/A')),
                        'ap_mac': str(row.get('AP MAC Address', 'N/A')),
                        'sm_mac': str(row.get('SM MAC Address', 'N/A'))
                    } for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))
                }
        for timestamp, sm_ip, status, latency in records:
            info = ip_info.get(sm_ip, {})
            logs.append({
                'time': timestamp,
                'entry': {
                    'ip': sm_ip,
                    'status': status,
                    'latency': latency,
                    'downtime_since': get_downtime_since(sm_ip, status),
                    'uptime_since': get_uptime_since(sm_ip, status),
                    'location': info.get('location', 'Unknown'),
                    'org_name': info.get('org_name', 'N/A'),
                    'ap_name': info.get('ap_name', 'N/A'),
                    'ap_ip': info.get('ap_ip', 'N/A'),
                    'ap_mac': info.get('ap_mac', 'N/A'),
                    'sm_mac': info.get('sm_mac', 'N/A')
                }
            })
        HISTORY_CACHE[cache_key] = {'data': logs, 'time': datetime.now()}
        logging.info(f"Retrieved {len(logs)} logs for IPs {ip_input}")
        return jsonify(logs)
    except Exception as e:
        logging.error(f"Get logs error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/ip_uptime_downtime')
def ip_uptime_downtime():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        ip_filter = request.args.get('ip_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            ip_info = {
                str(row['SM IP']): {
                    'location': str(row.get('Location', 'Unknown')),
                    'org_name': str(row.get('Device Name', 'N/A')),
                    'ap_name': str(row.get('AP Name', 'N/A'))
                } for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))
            }
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        ip_stats = defaultdict(lambda: {
            'uptime': 0,
            'downtime': 0,
            'degraded_time': 0,
            'down_count': 0,
            'up_count': 0,
            'down_events': [],
            'up_events': [],
            'degraded_events': []
        })
        for i, (sm_ip, status, timestamp) in enumerate(records):
            if ip_filter and sm_ip != ip_filter:
                continue
            ts = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            next_status = None
            next_ts = None
            if i + 1 < len(records) and records[i + 1][0] == sm_ip:
                next_status = records[i + 1][1]
                next_ts = datetime.strptime(records[i + 1][2], '%Y-%m-%d %H:%M:%S')
            duration = PING_INTERVAL
            if next_status and next_ts:
                duration = (next_ts - ts).total_seconds()
            if status == 'Down':
                ip_stats[sm_ip]['downtime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Down':
                    ip_stats[sm_ip]['down_count'] += 1
                    ip_stats[sm_ip]['down_events'].append(timestamp)
            elif status == 'Reachable':
                ip_stats[sm_ip]['uptime'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Reachable':
                    ip_stats[sm_ip]['up_count'] += 1
                    ip_stats[sm_ip]['up_events'].append(timestamp)
            elif status == 'Degraded':
                ip_stats[sm_ip]['degraded_time'] += duration
                if i == 0 or records[i-1][0] != sm_ip or records[i-1][1] != 'Degraded':
                    ip_stats[sm_ip]['degraded_events'].append(timestamp)
        result = {
            'ips': [],
            'chart_data': {
                'pie': {
                    'labels': ['Uptime', 'Downtime', 'Degraded'],
                    'data': []
                }
            }
        }
        for sm_ip in ip_stats:
            if ip_filter and sm_ip != ip_filter:
                continue
            info = ip_info.get(sm_ip, {})
            result['ips'].append({
                'sm_ip': sm_ip,
                'location': info.get('location', 'Unknown'),
                'org_name': info.get('org_name', 'N/A'),
                'ap_name': info.get('ap_name', 'N/A'),
                'uptime': format_duration(ip_stats[sm_ip]['uptime']),
                'downtime': format_duration(ip_stats[sm_ip]['downtime']),
                'degraded_time': format_duration(ip_stats[sm_ip]['degraded_time']),
                'down_count': ip_stats[sm_ip]['down_count'],
                'up_count': ip_stats[sm_ip]['up_count']
            })
            if sm_ip == ip_filter or not ip_filter:
                result['chart_data']['pie']['data'] = [
                    ip_stats[sm_ip]['uptime'],
                    ip_stats[sm_ip]['downtime'],
                    ip_stats[sm_ip]['degraded_time']
                ]
        logging.info(f"Retrieved uptime/downtime for {len(result['ips'])} IPs")
        return jsonify(result)
    except Exception as e:
        logging.error(f"IP uptime/downtime error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/location_health')
def location_health():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        location_filter = request.args.get('location_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            ip_to_location = {str(row['SM IP']): str(row['Location']) for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
            all_locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp, latency
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        health_data = defaultdict(lambda: {
            'total_latency': 0,
            'ping_count': 0,
            'down_count': 0,
            'reachable_count': 0,
            'degraded_count': 0,
            'ips': set()
        })
        for sm_ip, status, timestamp, latency in records:
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            health_data[location]['ips'].add(sm_ip)
            if status == 'Down':
                health_data[location]['down_count'] += 1
            elif status == 'Reachable' and latency is not None:
                health_data[location]['reachable_count'] += 1
                health_data[location]['total_latency'] += latency
                health_data[location]['ping_count'] += 1
            elif status == 'Degraded':
                health_data[location]['degraded_count'] += 1
        result = {
            'locations': [],
            'chart_data': {
                'bar': {
                    'labels': [],
                    'avg_latency': [],
                    'down_counts': []
                }
            }
        }
        locations_to_process = [location_filter] if location_filter else all_locations
        for location in locations_to_process:
            data = health_data.get(location, {
                'total_latency': 0,
                'ping_count': 0,
                'down_count': 0,
                'reachable_count': 0,
                'degraded_count': 0,
                'ips': set()
            })
            avg_latency = (data['total_latency'] / data['ping_count']) if data['ping_count'] > 0 else 0
            result['locations'].append({
                'location': location,
                'ip_count': len(data['ips']),
                'avg_latency': f"{avg_latency:.2f} ms" if avg_latency > 0 else 'N/A',
                'down_count': data['down_count'],
                'reachable_count': data['reachable_count'],
                'degraded_count': data['degraded_count']
            })
            result['chart_data']['bar']['labels'].append(location)
            result['chart_data']['bar']['avg_latency'].append(avg_latency)
            result['chart_data']['bar']['down_counts'].append(data['down_count'])
        logging.info(f"Retrieved health data for {len(result['locations'])} locations")
        return jsonify(result)
    except Exception as e:
        logging.error(f"Location health error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/export_health_pdf')
def export_health_pdf():
    if not session.get('logged_in'):
        return jsonify({'error': 'Unauthorized'}), 401
    try:
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        location_filter = request.args.get('location_filter')
        if not start_date or not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
            start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
        start_datetime = datetime.strptime(start_date, '%Y-%m-%d')
        end_datetime = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
        with results_lock:
            if CACHED_DF is None:
                return jsonify({'error': 'Dataframe not initialized'}), 500
            ip_to_location = {str(row['SM IP']): str(row['Location']) for row in CACHED_DF.to_dict('records') if pd.notna(row.get('SM IP'))}
            all_locations = sorted(set(CACHED_DF['Location'].dropna().astype(str)))
        with sqlite3.connect('ping_history.db', timeout=10) as conn:
            cursor = conn.cursor()
            query = """
                SELECT sm_ip, status, timestamp, latency
                FROM history
                WHERE timestamp >= ? AND timestamp < ?
                ORDER BY sm_ip, timestamp
            """
            cursor.execute(query, (start_datetime.strftime('%Y-%m-%d %H:%M:%S'), end_datetime.strftime('%Y-%m-%d %H:%M:%S')))
            records = cursor.fetchall()
        health_data = defaultdict(lambda: {
            'total_latency': 0,
            'ping_count': 0,
            'down_count': 0,
            'reachable_count': 0,
            'degraded_count': 0,
            'ips': set()
        })
        for sm_ip, status, timestamp, latency in records:
            location = ip_to_location.get(sm_ip, 'Unknown')
            if location_filter and location != location_filter:
                continue
            health_data[location]['ips'].add(sm_ip)
            if status == 'Down':
                health_data[location]['down_count'] += 1
            elif status == 'Reachable' and latency is not None:
                health_data[location]['reachable_count'] += 1
                health_data[location]['total_latency'] += latency
                health_data[location]['ping_count'] += 1
            elif status == 'Degraded':
                health_data[location]['degraded_count'] += 1
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(100, 750, "Location Health Report")
        c.setFont("Helvetica", 12)
        c.drawString(100, 730, f"Period: {start_date} to {end_date}")
        y = 700
        c.setFont("Helvetica-Bold", 12)
        c.drawString(50, y, "Location")
        c.drawString(150, y, "IP Count")
        c.drawString(250, y, "Avg Latency")
        c.drawString(350, y, "Down Count")
        c.drawString(450, y, "Reachable Count")
        y -= 20
        c.setFont("Helvetica", 10)
        locations_to_process = [location_filter] if location_filter else all_locations
        for location in locations_to_process:
            data = health_data.get(location, {
                'total_latency': 0,
                'ping_count': 0,
                'down_count': 0,
                'reachable_count': 0,
                'degraded_count': 0,
                'ips': set()
            })
            avg_latency = (data['total_latency'] / data['ping_count']) if data['ping_count'] > 0 else 0
            c.drawString(50, y, location)
            c.drawString(150, y, str(len(data['ips'])))
            c.drawString(250, y, f"{avg_latency:.2f} ms" if avg_latency > 0 else 'N/A')
            c.drawString(350, y, str(data['down_count']))
            c.drawString(450, y, str(data['reachable_count']))
            y -= 20
            if y < 50:
                c.showPage()
                y = 750
        c.save()
        buffer.seek(0)
        logging.info(f"Exported health PDF for {start_date} to {end_date}")
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"location_health_{start_date}_to_{end_date}.pdf"
        )
    except Exception as e:
        logging.error(f"Export health PDF error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/')
def index():
    return render_template('index.html')

if __name__ == '__main__':
    start_periodic_update()
    socketio.run(app, host='0.0.0.0', port=5000, debug=False)
